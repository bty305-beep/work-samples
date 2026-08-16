#!/usr/bin/env python3
"""
商品マスタと除外リストから、販売リストをボタン一つ（1コマンド）で生成します。

    python generate.py                  # config.json を読んで実行
    python generate.py config.json      # 設定ファイルを指定して実行

やっていること。

    1. 商品マスタ（商品名・仕入れ値・商品コード）を読む
    2. 除外リスト（商品コード）に載っているものを落とす
       … このとき商品コードの表記ゆれ（全角・空白・大文字小文字）を吸収します
    3. 設定ファイルの計算ルールで販売価格を決める（掛け率・端数処理・下限価格）
    4. 販売リストのExcelと、処理サマリーを書き出す

数十万件を想定しているため、行ごとのループは使わず pandas でまとめて処理します。
Excelの書き出しも1行ずつメモリに溜めない方式（constant_memory）です。
"""

import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

HERE = Path(__file__).resolve().parent

# 商品マスタから受け取る項目。設定ファイルでは、この名前に対して
# 「お使いのファイルでの列名」を割り当てます。
CODE, NAME, COST = "商品コード", "商品名", "仕入れ値"

# 出力できる項目と、Excelでの表示形式。
NUMBER_FORMATS = {
    "仕入れ値": "#,##0",
    "販売価格": "#,##0",
    "粗利": "#,##0",
    "粗利率": "0.0%",
    "掛け率": "0.00",
}

EXCEL_MAX_ROWS = 1_048_576


# --------------------------------------------------------------------------
# 商品コードの正規化
#
# 除外したい商品が除外されない事故は、ほぼここで起きます。
# 「ＡＢ－0012」と「ab-0012」と「AB-0012 」は、人間には同じコードですが
# 文字列としては別物です。突合の前に、必ず同じ形に揃えます。
# --------------------------------------------------------------------------

def normalize_key(series, rules):
    """商品コードを突合用の形に揃えます。元の表記は残したまま、別列で持ちます。"""
    s = series.astype("string").fillna("")

    if rules.get("nfkc", True):
        # 全角英数字・全角記号・全角スペースを半角に寄せます（ＡＢ－０１ → AB-01）
        s = s.str.normalize("NFKC")
    if rules.get("strip", True):
        s = s.str.strip()
    if rules.get("strip_spaces", True):
        # 途中に入り込んだ空白も落とします（AB - 01 → AB-01）
        s = s.str.replace(r"\s+", "", regex=True)
    if rules.get("upper", True):
        s = s.str.upper()
    if rules.get("remove_chars"):
        # ハイフンの有無がゆれる運用なら "-" を指定します（AB-01 と AB01 を同一視）
        s = s.str.replace(f"[{re.escape(rules['remove_chars'])}]", "", regex=True)
    if rules.get("digits_only"):
        s = s.str.replace(r"\D", "", regex=True)

    return s


# --------------------------------------------------------------------------
# 読み込み
# --------------------------------------------------------------------------

def read_table(spec, required):
    """Excel でも CSV でも同じように読みます。商品コードは必ず文字列として読みます。

    商品コードを数値として読ませてはいけません。`0012` が `12` になり、
    `1.0E+05` のような表記に化けることもあります。除外漏れの典型的な原因です。
    """
    path = Path(spec["path"])
    if not path.is_absolute():
        path = HERE / path
    if not path.exists():
        sys.exit(f"ファイルが見つかりません: {path}")

    columns = spec.get("columns", {})
    actual = {key: columns.get(key, key) for key in required}
    dtype = {actual[CODE]: str} if CODE in required else None

    if path.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        df = pd.read_excel(
            path,
            sheet_name=spec.get("sheet", 0),
            header=spec.get("header_row", 0),
            dtype=dtype,
        )
    else:
        df = pd.read_csv(
            path,
            encoding=spec.get("encoding", "utf-8-sig"),
            header=spec.get("header_row", 0),
            dtype=dtype,
        )

    df.columns = [str(c).strip() for c in df.columns]

    missing = [actual[key] for key in required if actual[key] not in df.columns]
    if missing:
        sys.exit(
            f"{path.name} に次の列がありません: {' / '.join(missing)}\n"
            f"  このファイルにある列: {' / '.join(map(str, df.columns))}\n"
            f"  設定ファイルの columns を、お使いのファイルの列名に合わせてください。"
        )

    # 必要な列だけに絞ります。数十万件のとき、余計な列を持ち回らないだけで軽くなります。
    df = df[[actual[key] for key in required]]
    df.columns = list(required)
    print(f"  読み込み: {path.name} → {len(df):,}行")
    return df


# --------------------------------------------------------------------------
# 仕入れ値の数値化
# --------------------------------------------------------------------------

def to_amount(series, clean=True):
    """仕入れ値を数値にします。「1,200円」「１２００」のような表記も拾います。"""
    values = pd.to_numeric(series, errors="coerce").astype("float64")
    if not clean:
        return values

    # 数値として読めなかった行だけを文字列として整形し、もう一度試します。
    # 全件に文字列処理をかけると数十万件では無駄が大きいので、失敗分だけ処理します。
    broken = values.isna() & series.notna()
    if broken.any():
        text = (
            series[broken]
            .astype("string")
            .str.normalize("NFKC")
            .str.replace(r"[^\d.\-]", "", regex=True)
            .replace("", pd.NA)
        )
        recovered = pd.to_numeric(text, errors="coerce")
        values.loc[broken] = recovered.to_numpy(dtype="float64", na_value=np.nan)
    return values


# --------------------------------------------------------------------------
# 販売価格の計算
#
# 計算ルールはすべて設定ファイル側にあります。Pythonのコードは触りません。
# --------------------------------------------------------------------------

def markup_of(cost, pricing):
    """掛け率を返します。tiers があれば仕入れ値の帯ごとに掛け率を変えます。"""
    tiers = pricing.get("tiers")
    if not tiers:
        return float(pricing.get("markup", 1.0))

    # 「1000円まで1.8倍、5000円まで1.5倍、それ以上1.3倍」のような段階設定。
    # searchsorted で一括判定します（行ループなし）。
    bounds, rates = [], []
    for tier in tiers:
        upto = tier.get("upto")
        bounds.append(np.inf if upto is None else float(upto))
        rates.append(float(tier["markup"]))
    order = np.argsort(bounds)
    bounds = np.asarray(bounds)[order]
    rates = np.asarray(rates)[order]

    index = np.searchsorted(bounds, cost.to_numpy(dtype=float), side="left")
    index = np.clip(index, 0, len(rates) - 1)
    return pd.Series(rates[index], index=cost.index)


def apply_round(price, rule):
    """端数処理。単位（10円単位など）と、切り上げ／切り捨て／四捨五入を指定します。"""
    if not rule:
        return price
    unit = float(rule.get("unit", 1))
    if unit <= 0:
        return price

    mode = rule.get("mode", "ceil")
    scaled = price / unit
    if mode == "ceil":
        scaled = np.ceil(scaled)
    elif mode == "floor":
        scaled = np.floor(scaled)
    elif mode == "round":
        # numpy の round は偶数丸めのため、日本の四捨五入になるよう明示します。
        scaled = np.floor(scaled + 0.5)
    else:
        sys.exit(f"round.mode には ceil / floor / round のいずれかを指定してください: {mode}")
    return scaled * unit


def compute_price(cost, pricing):
    """仕入れ値から販売価格を作ります。戻り値は 販売価格 と 掛け率。"""
    markup = markup_of(cost, pricing)
    price = cost * markup + float(pricing.get("add", 0))
    price = apply_round(price, pricing.get("round"))

    floor_price = pricing.get("min_price")
    hit_floor = 0
    if floor_price is not None:
        below = price < float(floor_price)
        hit_floor = int(below.sum())
        price = price.where(~below, float(floor_price))

    return price, markup, hit_floor


# --------------------------------------------------------------------------
# 書き出し
# --------------------------------------------------------------------------

def pick_engine(preferred):
    if preferred in ("xlsxwriter", "openpyxl"):
        return preferred
    try:
        import xlsxwriter  # noqa: F401
        return "xlsxwriter"
    except ImportError:
        return "openpyxl"


def split_sheets(name, df):
    """Excelの上限（約104万行）を超える場合はシートを分けます。"""
    limit = EXCEL_MAX_ROWS - 1
    if len(df) <= limit:
        return [(name, df)]
    parts = []
    for i in range(0, len(df), limit):
        parts.append((f"{name}_{i // limit + 1}", df.iloc[i:i + limit]))
    return parts


def prepare_column(series):
    """Excelに渡せる値に整えます。空欄は空文字、numpy の数値は素の数値にします。"""
    values = series.tolist()
    if series.dtype == object or bool(series.isna().any()):
        cleaned = []
        for value in values:
            if value is None or value is pd.NA:
                cleaned.append("")
            elif isinstance(value, float) and value != value:  # NaN
                cleaned.append("")
            elif isinstance(value, np.generic):
                cleaned.append(value.item())
            else:
                cleaned.append(value)
        return cleaned
    return values


def column_width(header, sample):
    widths = [len(str(header))] + [len(str(v)) for v in sample]
    return min(max(widths) * 1.4 + 4, 40)


def write_workbook(path, sheets, engine="auto"):
    """販売リストを書き出します。1行ずつ流し込むため、件数が増えてもメモリが膨らみません。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    engine = pick_engine(engine)

    pages = []
    for name, df in sheets:
        pages.extend(split_sheets(name, df))

    if engine == "xlsxwriter":
        import xlsxwriter

        book = xlsxwriter.Workbook(str(path), {"constant_memory": True})
        header_format = book.add_format({"bold": True, "bg_color": "#EFEFEF", "align": "center"})
        formats = {}
        for name, df in pages:
            sheet = book.add_worksheet(name)
            columns = list(df.columns)
            sheet.write_row(0, 0, columns, header_format)
            sheet.freeze_panes(1, 0)

            data = [prepare_column(df[c]) for c in columns]
            for i, column in enumerate(columns):
                number_format = NUMBER_FORMATS.get(column)
                if number_format and number_format not in formats:
                    formats[number_format] = book.add_format({"num_format": number_format})
                sheet.set_column(
                    i, i,
                    column_width(column, data[i][:200]),
                    formats.get(number_format),
                )

            for row in range(len(df)):
                sheet.write_row(row + 1, 0, [values[row] for values in data])
        book.close()
    else:
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        # write_only モードは、行を書いたそばから吐き出すのでメモリを持ちません。
        book = Workbook(write_only=True)
        for name, df in pages:
            sheet = book.create_sheet(name)
            columns = list(df.columns)
            data = [prepare_column(df[c]) for c in columns]

            header = []
            for i, column in enumerate(columns):
                sheet.column_dimensions[get_column_letter(i + 1)].width = column_width(
                    column, data[i][:200]
                )
                cell = WriteOnlyCell(sheet, value=column)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="EFEFEF")
                cell.alignment = Alignment(horizontal="center")
                header.append(cell)
            sheet.append(header)

            for row in range(len(df)):
                sheet.append([values[row] for values in data])
        book.save(path)

    return engine


# --------------------------------------------------------------------------

def build(config, started):
    stats = {}
    timings = {}

    # --- 読み込み ---------------------------------------------------------
    clock = time.perf_counter()
    master = read_table(config["master"], [CODE, NAME, COST])
    excludes = read_table(config["exclude"], [CODE])
    timings["読み込み"] = time.perf_counter() - clock

    stats["入力件数"] = len(master)
    stats["除外リスト件数"] = len(excludes)

    # --- 突合と計算 -------------------------------------------------------
    clock = time.perf_counter()
    rules = config.get("code_rules", {})
    master["_key"] = normalize_key(master[CODE], rules)
    exclude_keys = normalize_key(excludes[CODE], rules)
    exclude_keys = exclude_keys[exclude_keys != ""]

    unique_keys = pd.Index(exclude_keys.unique())
    stats["除外コード（重複を除く）"] = len(unique_keys)
    stats["除外リストにあるがマスタに無いコード"] = int((~unique_keys.isin(master["_key"])).sum())

    dropped = []  # 除外した行を、理由つきで控えておきます

    # マスタ内の重複コード
    duplicated = master["_key"].duplicated(keep="first")
    stats["マスタの重複コード"] = int(duplicated.sum())
    if config.get("duplicate_codes", "first") == "first" and duplicated.any():
        dropped.append(master[duplicated].assign(除外理由="コード重複"))
        master = master[~duplicated]

    # 除外リストとの突合（ここが本題）
    hit = master["_key"].isin(unique_keys)
    stats["除外件数"] = int(hit.sum())
    if hit.any():
        dropped.append(master[hit].assign(除外理由="除外リスト該当"))
    master = master[~hit]

    # 仕入れ値
    master[COST] = to_amount(master[COST], clean=config.get("pricing", {}).get("clean_cost", True))
    broken = master[COST].isna()
    stats["仕入れ値が読めない行"] = int(broken.sum())
    if broken.any():
        dropped.append(master[broken].assign(除外理由="仕入れ値が数値でない"))
    master = master[~broken]

    pricing = config.get("pricing", {})
    price, markup, hit_floor = compute_price(master[COST], pricing)
    master["販売価格"] = price
    master["掛け率"] = markup
    master["粗利"] = master["販売価格"] - master[COST]
    master["粗利率"] = (master["粗利"] / master["販売価格"]).round(4)
    if pricing.get("min_price") is not None:
        stats["下限価格まで引き上げた件数"] = hit_floor

    stats["出力件数"] = len(master)
    timings["突合と計算"] = time.perf_counter() - clock

    # --- 書き出し ---------------------------------------------------------
    clock = time.perf_counter()
    output = config["output"]
    columns = output.get("columns", [CODE, NAME, COST, "販売価格", "粗利", "粗利率"])
    unknown = [c for c in columns if c not in master.columns]
    if unknown:
        sys.exit(
            f"output.columns に使えない項目があります: {' / '.join(unknown)}\n"
            f"  使える項目: {' / '.join([CODE, NAME, COST, '販売価格', '掛け率', '粗利', '粗利率'])}"
        )

    result = master[columns].sort_values(output.get("sort_by", CODE), kind="stable")

    stamp = started.strftime(output.get("timestamp_format", "%Y%m%d_%H%M"))
    path = Path(output["path"].replace("{日時}", stamp).replace("{timestamp}", stamp))
    if not path.is_absolute():
        path = HERE / path

    sheets = [(output.get("sheet", "販売リスト"), result)]
    if output.get("include_excluded_sheet", True) and dropped:
        excluded = pd.concat(dropped, ignore_index=True)
        sheets.append(("除外一覧", excluded[[CODE, NAME, COST, "除外理由"]]))

    engine = write_workbook(path, sheets, output.get("engine", "auto"))
    timings["Excel書き出し"] = time.perf_counter() - clock

    csv_path = None
    if output.get("csv"):
        clock = time.perf_counter()
        csv_path = Path(output["csv"].replace("{日時}", stamp).replace("{timestamp}", stamp))
        if not csv_path.is_absolute():
            csv_path = HERE / csv_path
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        result.to_csv(csv_path, index=False, encoding="utf-8-sig")
        timings["CSV書き出し"] = time.perf_counter() - clock

    return path, csv_path, stats, timings, engine


def main():
    config_path = Path(sys.argv[1]) if len(sys.argv) > 1 else HERE / "config.json"
    if not config_path.is_absolute():
        config_path = HERE / config_path
    if not config_path.exists():
        sys.exit(f"設定ファイルが見つかりません: {config_path}")

    config = json.loads(config_path.read_text(encoding="utf-8"))
    job_name = config.get("job_name", "販売リスト生成")

    print(f"{job_name} を実行します。")
    started = datetime.now()
    clock = time.perf_counter()
    path, csv_path, stats, timings, engine = build(config, started)
    elapsed = time.perf_counter() - clock
    stats["処理秒数"] = round(elapsed, 2)

    print()
    for key, value in stats.items():
        print(f"  {key}: {value:,}" if isinstance(value, int) else f"  {key}: {value}")
    print()
    for key, value in timings.items():
        print(f"  {key}: {value:.2f}秒")
    print(f"  合計: {elapsed:.2f}秒（書き出しエンジン: {engine}）")
    print()
    print(f"書き出しました: {path}")
    if csv_path:
        print(f"書き出しました: {csv_path}")

    if config.get("summary"):
        summary_path = Path(config["summary"])
        if not summary_path.is_absolute():
            summary_path = HERE / summary_path
        lines = [
            f"# {job_name}",
            "",
            f"実行日時: {started.strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "## 件数",
            "",
            "| 項目 | 件数 |",
            "|---|---:|",
        ]
        for key, value in stats.items():
            if key == "処理秒数":
                continue
            lines.append(f"| {key} | {value:,} |")
        lines += [
            "",
            "## 処理時間",
            "",
            "| 工程 | 秒 |",
            "|---|---:|",
        ]
        for key, value in timings.items():
            lines.append(f"| {key} | {value:.2f} |")
        lines.append(f"| **合計** | **{elapsed:.2f}** |")
        lines += ["", f"出力: `{path.name}`"]
        if csv_path:
            lines.append(f"出力: `{csv_path.name}`")
        summary_path.parent.mkdir(parents=True, exist_ok=True)
        summary_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        print(f"サマリーを書き出しました: {summary_path}")


if __name__ == "__main__":
    main()
