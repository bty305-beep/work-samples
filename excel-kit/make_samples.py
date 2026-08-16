#!/usr/bin/env python3
"""
動作確認用のサンプルExcelを作ります。

Excelはバイナリのためリポジトリに置かず、このスクリプトで生成する形にしています。

    python make_samples.py

samples/ に5ファイルを書き出します。

    請求データ_2026-08.xlsx    請求番号・取引先・請求日・請求金額
    入金データ_2026-08.xlsx    振込依頼人・入金日・入金額
    売上_本店.xlsx             日付・メニュー・金額
    売上_北営業所.xlsx         〃
    売上_南工場.xlsx           〃

入金データには、突合デモとして次のずれを意図的に入れてあります。

    振込手数料440円が引かれている        → 金額不一致（差 440）
    一部入金で50,000円足りない           → 金額不一致（差 50,000）
    請求番号ではなく取引先名で振り込まれた → 請求のみ／入金のみ に分かれる
    請求側に存在しない入金（前月分）      → 入金のみ
    請求番号の大文字小文字がゆれている    → 設定で吸収して一致する
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SAMPLES = Path(__file__).resolve().parent / "samples"

# 作成日時を固定しておくと、作り直しても中身が変わりません。
CREATED = datetime(2026, 8, 1)

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="EFEFEF")


# --------------------------------------------------------------------------
# 請求と入金（reconcile 用）
# --------------------------------------------------------------------------

# 請求番号 / 取引先 / 請求日 / 請求金額
INVOICES = [
    ("INV-2026-0801", "株式会社アオゾラ物流", "2026-08-03", 286000),
    ("INV-2026-0802", "ひまわり工業株式会社", "2026-08-05", 154000),
    ("INV-2026-0803", "有限会社ミドリ商会",   "2026-08-07", 198000),
    ("INV-2026-0804", "株式会社しおかぜ設計", "2026-08-11", 462000),
    ("INV-2026-0805", "はやて運輸株式会社",   "2026-08-17",  93500),
    ("INV-2026-0806", "株式会社こもれび食品", "2026-08-21", 220000),
    ("INV-2026-0807", "やまびこ印刷株式会社", "2026-08-25", 176000),
]

# 振込依頼人 / 入金日 / 入金額
# 振込依頼人には通常このように請求番号を書いてもらいますが、現場では必ずずれます。
PAYMENTS = [
    ("inv-2026-0801",      "2026-08-14", 286000),  # 大文字小文字のゆれ → 設定で吸収して一致
    ("INV-2026-0802",      "2026-08-20", 154000),  # 一致
    ("INV-2026-0804",      "2026-08-28", 412000),  # 一部入金 → 差 50,000
    ("INV-2026-0805",      "2026-08-31",  93060),  # 振込手数料440円が引かれている → 差 440
    ("INV-2026-0807",      "2026-08-26", 176000),  # 一致
    ("有限会社ミドリ商会", "2026-08-25", 198000),  # 請求番号ではなく取引先名で振り込まれた
    ("INV-2026-0731",      "2026-08-05",  88000),  # 前月分。今回の請求側には存在しない
]


# --------------------------------------------------------------------------
# 拠点別の売上（aggregate 用）
# --------------------------------------------------------------------------

# 日付 / メニュー / 金額
SALES = {
    "本店": [
        ("2026-08-03", "定期メンテナンス", 48000),
        ("2026-08-05", "スポット対応",     22000),
        ("2026-08-07", "部材販売",         13500),
        ("2026-08-11", "設置工事",        180000),
        ("2026-08-14", "定期メンテナンス", 48000),
        ("2026-08-19", "スポット対応",     35000),
        ("2026-08-24", "部材販売",          8200),
        ("2026-08-28", "定期メンテナンス", 52000),
    ],
    "北営業所": [
        ("2026-08-04", "定期メンテナンス", 36000),
        ("2026-08-06", "設置工事",        124000),
        ("2026-08-12", "スポット対応",     18000),
        ("2026-08-18", "定期メンテナンス", 36000),
        ("2026-08-20", "部材販売",         26400),
        ("2026-08-25", "スポット対応",     19500),
        ("2026-08-31", "設置工事",         96000),
    ],
    "南工場": [
        ("2026-08-05", "設置工事",        210000),
        ("2026-08-08", "部材販売",         42000),
        ("2026-08-13", "定期メンテナンス", 64000),
        ("2026-08-21", "設置工事",        165000),
        ("2026-08-26", "部材販売",         31000),
        ("2026-08-29", "定期メンテナンス", 64000),
    ],
}


# --------------------------------------------------------------------------

def write_sheet(path, title, headers, rows, money_columns=()):
    """見出しを固定し、金額列に桁区切りを付けた状態で書き出します。"""
    wb = Workbook()
    wb.properties.created = CREATED
    wb.properties.modified = CREATED
    wb.properties.creator = "make_samples.py"
    ws = wb.active
    ws.title = title

    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(list(row))

    for i, header in enumerate(headers, 1):
        letter = get_column_letter(i)
        width = max([len(str(header))] + [len(str(r[i - 1])) for r in rows]) + 4
        ws.column_dimensions[letter].width = min(width * 1.4, 30)
        if header in money_columns:
            for cell in ws[letter][1:]:
                cell.number_format = "#,##0"

    ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  {path.name}: {len(rows)}行")
    return path


def main():
    print(f"サンプルを作成します: {SAMPLES}")

    write_sheet(
        SAMPLES / "請求データ_2026-08.xlsx", "請求",
        ["請求番号", "取引先", "請求日", "請求金額"], INVOICES,
        money_columns={"請求金額"},
    )
    write_sheet(
        SAMPLES / "入金データ_2026-08.xlsx", "入金",
        ["振込依頼人", "入金日", "入金額"], PAYMENTS,
        money_columns={"入金額"},
    )
    for name, rows in SALES.items():
        write_sheet(
            SAMPLES / f"売上_{name}.xlsx", name,
            ["日付", "メニュー", "金額"], rows,
            money_columns={"金額"},
        )

    print()
    print("できました。続けて次を実行すると、突合・集計・転記が動きます。")
    print("  python run.py jobs/reconcile.json")
    print("  python run.py jobs/aggregate.json")
    print("  python run.py jobs/fill.json")


if __name__ == "__main__":
    main()
